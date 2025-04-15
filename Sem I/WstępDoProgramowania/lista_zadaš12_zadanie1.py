"""
    Wdp labolatorium
    Barbara Dereń 279914
    L12
"""

import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.utils import get_column_letter
# Zadanie 1

# Wczytanie pliku XLSX przy użyciu openpyxl
nazwa_pliku = '/Users/vusis/Documents/Program/WstępDoProgramowania/L11 g1/dane_g1.xlsx'
arkusz = 'TABLICA'
wb = openpyxl.load_workbook(nazwa_pliku)
sheet = wb[arkusz]

# Znalezienie komórek z nagłówkami
wskaznik = []  # lista wskaznikow z ktorych dane maja zostac uwzgldnione
lata = []  # lista list z latami, ktore zostaną uwzględnione w danych
wojewodztwa = []
wiersze_z_wojewodztwami = []
dane_lista = []
dane_slownik = {}
start_column = 'C'
end_column = chr(ord(start_column) + 8)  # Kolumna na którą kończy się 'zestaw' kolumn pierwszego wskaznika

for cell in sheet['B']:  # iterowanie po kolejnych wojewodztwach
    if cell.font.bold:  # sprawdzanie czy są one pogrubione
        wojewodztwa.append(cell.value)
        wiersze_z_wojewodztwami.append(cell.row)

if wojewodztwa != []:  # warunek - muszą być wybrane wojewodztwa
    while True:
        for column in sheet.iter_cols(min_col=openpyxl.utils.column_index_from_string(start_column),
                                      max_col=openpyxl.utils.column_index_from_string(end_column)):
            nazwa_wskaznika = ''
            lata_pom = []
            dane_pom = []

            for wojewodztwo in wojewodztwa:  # 'czyszczenie' zestawu danych w slowniku
                dane_slownik[wojewodztwo] = []  # stworzenie kluczy w slowniku jako wojewodztwa

            if column[1].font.bold:  # sprawdzanie czy wskaźnik jest pogrubiony

                for cell in sheet['3']:
                    if cell.column >= openpyxl.utils.column_index_from_string(start_column) and \
                            cell.column <= openpyxl.utils.column_index_from_string(end_column) and \
                            cell.font.bold:  # sprawdzanie czy komórki w 3 wierszu (a więc lata) znajdują się w odpowiednim przedziale kolumn oraz czy są pogrubiobe

                        lata_pom.append(cell.value)
                        nazwa_wskaznika = column[1].value

                        for dana in sheet[get_column_letter(
                                cell.column)]:  # kolejne komorki w kolumnach w ktorych lata byly pogrubione
                            for wiersz in wiersze_z_wojewodztwami:
                                if dana.row == wiersz:
                                    dane_pom.append(dana.value)

                if nazwa_wskaznika != '':
                    wskaznik.append(nazwa_wskaznika)
                    lata.append(lata_pom)
                    for i in range(len(wojewodztwa)):
                        for j in range(0, len(dane_pom), len(wojewodztwa)):
                            dane_slownik[wojewodztwa[i]].append(
                                dane_pom[i + j])  # aktualizacja slownika zawierajacego dane
                    dane_lista.append(dane_slownik.copy())  # aktualizacja listy slownikow zawierajacych wszystkie dane

        start_column = get_column_letter(
            openpyxl.utils.column_index_from_string(
                end_column) + 1)  # Przesunięcie o kolumne do kolejnego zestawu 9 kolumn
        end_column = get_column_letter(
            openpyxl.utils.column_index_from_string(start_column) + 8)  # Kolumna na którą kończy się kolejnego zestawu

        if not sheet[start_column + '2'].value:  # Sprawdzenie czy jest kolejny zestaw
            break

    print(dane_slownik, wskaznik, lata)
    if lata != [] and wskaznik != []:  # warunek istnienia zestawów wskaznika i lat, aby można było narysować wykres
        fig, axs = plt.subplots(len(wskaznik), 1, figsize=(4, 3 * len(wskaznik)))  # Tworzenie nowej figury
        style = ['/', '.', '-', '*'] #ponieważ wykres jest slupkowy nie zawiera on linii wiec, zeby nadac roznym seriom danych rozne style uzyje roznych wzorow ich wypelnienia
        for i in range(len(wskaznik)):
            bar_width = 0.1
            index = np.arange(len(lata[i]))

            for j, wojewodztwo in enumerate(
                    wojewodztwa):  # Tworzenie wykresów kolumnowych dla kolejnych zestawów danych z pogrupowanymi słupkami dla kolejnych lat
                axs[i].bar(index + j * bar_width, dane_lista[i][wojewodztwo], bar_width,
                           label=wojewodztwo, hatch=style[j])  # wczytanie danych do wykresu

            axs[i].set_xlabel('Lata')
            axs[i].set_ylabel('Wartości [%]')
            axs[i].set_title(wskaznik[i])
            axs[i].set_xticks(index + bar_width * len(wojewodztwa) / 2, lata[i])  # dostowanie etykiet na osi poziomej
            axs[i].legend()

        plt.tight_layout()
        plt.show()
    else:
        print('Brak danych')
else:
    print('Brak danych')
